import streamlit as st
import pandas as pd
import re, math
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment

# =========================
# KONFIGURASI
# =========================
TEMPLATE_PATH = "templates/template_surat.xlsx"
START_TABLE_ROW = 19

st.title("Ekstraksi Surat Desa - Excel Automation")

# =========================
# UTIL
# =========================
def rupiah_format(cell):
    try:
        cell.value = float(str(cell.value).replace(".", "").replace(",", ""))
    except:
        pass
    cell.number_format = '#,##0'


def format_nama_desa(text):
    return text.title() if text else text


# =========================
# EKSTRAK INFO SURAT
# =========================
def extract_info(df):
    nama_desa = tanggal = nomor = None

    for r in range(len(df)):
        for c in range(len(df.columns)):
            val = str(df.iat[r, c]) if pd.notna(df.iat[r, c]) else ""

            if not nama_desa and "DESA" in val.upper():
                nama_desa = val.strip()

            if not tanggal:
                m = re.search(r"\d{1,2}\s\w+\s20\d{2}", val)
                if m:
                    tanggal = m.group()

            if not nomor and "NOMOR" in val.upper():
                try:
                    nomor = str(df.iat[r, c+2]).strip()
                except:
                    pass

    return format_nama_desa(nama_desa), tanggal, nomor


# =========================
# EKSTRAK TABEL SPM
# =========================
def extract_table(df, uploaded):
    header_row = None

    # 1. Cari header tabel
    for i in range(len(df)):
        row_text = " ".join(str(x).upper() for x in df.iloc[i].fillna(""))
        if "NO" in row_text and "SPM" in row_text and "ANGGARAN" in row_text:
            header_row = i
            break

    if header_row is None:
        return pd.DataFrame()

    # 2. Baca tabel mulai header
    table = pd.read_excel(uploaded, header=header_row)
    table.columns = [str(c).upper() for c in table.columns]

    # 3. Mapping kolom WAJIB
    col_no = next(c for c in table.columns if c.endswith("NO"))
    col_spm = next(c for c in table.columns if "SPM" in c)
    col_kegiatan = next(c for c in table.columns if "KEGIATAN" in c)
    col_anggaran = next(c for c in table.columns if "ANGGARAN" in c)
    col_ket = next(c for c in table.columns if "KET" in c)

    table = table[[col_no, col_spm, col_kegiatan, col_anggaran, col_ket]]
    table.columns = ["NO", "NOMOR SPM", "KEGIATAN", "ANGGARAN", "KETERANGAN"]

    # 4. FILTER KETAT: stop saat NO bukan angka
    cleaned = []

    for _, row in table.iterrows():
        no_val = str(row["NO"]).strip()

        # STOP TOTAL: begitu NO bukan angka
        if not no_val.isdigit():
            break

        cleaned.append(row)

    return pd.DataFrame(cleaned)


# =========================
# TULIS TABEL
# =========================
def write_table(ws, table_df):
    thin = Border(*(Side(style='thin') for _ in range(4)))
    row = START_TABLE_ROW

    for _, data in table_df.iterrows():
        ws.cell(row, 3).value = data["NO"]
        ws.cell(row, 4).value = data["NOMOR SPM"]
        ws.cell(row, 5).value = data["KEGIATAN"]
        ws.cell(row, 7).value = data["ANGGARAN"]
        ws.cell(row, 8).value = data["KETERANGAN"]

        ws.merge_cells(
            start_row=row,
            start_column=5,
            end_row=row,
            end_column=6
        )


        for col in [3,4,5,6,7,8]:
            ws.cell(row, col).border = thin

        row += 1

    return row


# =========================
# BARIS JUMLAH
# =========================
def write_total(ws, table_end_row):
    thin = Border(*(Side(style='thin') for _ in range(4)))

    ws.merge_cells(
        start_row=table_end_row,
        start_column=3,
        end_row=table_end_row,
        end_column=6
    )

    ws.cell(table_end_row, 3).value = "JUMLAH"
    ws.cell(table_end_row, 3).font = Font(bold=True)
    ws.cell(table_end_row, 3).alignment = Alignment(horizontal="center")

    total_cell = ws.cell(
        table_end_row,
        7,
        f"=SUM(G{START_TABLE_ROW+1}:G{table_end_row-1})"
    )

    total_cell.number_format = '#,##0'
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal="right", vertical="center")

    for col in range(3, 9):
        ws.cell(table_end_row, col).border = thin

    return table_end_row + 1



def calc_row_height(ws, text, start_col, end_col, base_height=15):
    """
    Hitung tinggi baris berdasarkan lebar kolom merged (Excel-like)
    """
    if not text:
        return base_height

    # total lebar kolom merged (approx)
    total_width = 0
    for col in range(start_col, end_col + 1):
        col_letter = chr(64 + col)  # 5->E, 6->F
        width = ws.column_dimensions[col_letter].width or 10
        total_width += width

    # estimasi karakter per baris (empiris)
    chars_per_line = int(total_width * 1.1)

    lines = math.ceil(len(text) / max(chars_per_line, 1))
    return max(base_height, lines * base_height)


# =========================
# FORMAT TABEL
# =========================
def format_table(ws, table_end_row):
    for r in range(START_TABLE_ROW, table_end_row):

        # =========================
        # BARIS PERTAMA SETELAH HEADER (NOMOR KOLOM)
        # =========================
        if r == START_TABLE_ROW:
            for col in [3, 4, 5, 6, 7, 8]:
                ws.cell(r, col).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
            continue   # <-- PENTING: jangan lanjut format baris biasa

        # =========================
        # BARIS DATA BIASA
        # =========================

        # NO
        ws.cell(r,3).alignment = Alignment(horizontal="center", vertical="center")

        # NOMOR SPM
        ws.cell(r,4).alignment = Alignment(horizontal="left", vertical="center")

        # KEGIATAN (MERGE + WRAP + AUTO HEIGHT)
        ws.merge_cells(
            start_row=r,
            start_column=5,
            end_row=r,
            end_column=6
        )

        cell = ws.cell(r,5)
        text = str(cell.value or "")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        row_height = calc_row_height(ws, text, start_col=5, end_col=6)
        ws.row_dimensions[r].height = row_height


        # ANGGRAN
        rupiah_format(ws.cell(r,7))
        ws.cell(r,7).alignment = Alignment(horizontal="right", vertical="center")

        # KETERANGAN
        ws.cell(r,8).alignment = Alignment(horizontal="center", vertical="center")



# =========================
# TANDA TANGAN
# =========================
def write_signature(ws, start_row):
    # start_row = baris setelah JUMLAH
    r = start_row + 1   # 1 baris kosong setelah tabel

    def center_merge(row, text, bold=False, underline=False):
        ws.merge_cells(
            start_row=row,
            start_column=5,
            end_row=row,
            end_column=8
        )
        ws.cell(row, 5).value = text
        ws.cell(row, 5).alignment = Alignment(horizontal="center")
        ws.cell(row, 5).font = Font(
            bold=bold,
            underline="single" if underline else None
        )

    # =========================
    # TEKS PENUTUP
    # =========================
    ws.merge_cells(
        start_row=r,
        start_column=3,
        end_row=r,
        end_column=8
    )
    ws.cell(r, 3).value = "Demikian untuk menjadikan periksa."
    ws.cell(r, 3).alignment = Alignment(horizontal="left")

    # =========================
    # TANDA TANGAN CAMAT
    # =========================
    r = r + 2   # 1 baris kosong setelah teks penutup

    center_merge(r, "Plt. CAMAT GRABAGAN", bold=True)
    center_merge(r + 4, "AGUS HERU PURNOMO, SP", bold=True, underline=True)
    center_merge(r + 5, "Pembina")
    center_merge(r + 6, "NIP 19690701 199703 1 008")



# =========================
# ISI TEMPLATE
# =========================
def fill_to_template(info_df, table_df):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    ws["F6"] = f"Grabagan, {info_df.loc[1,'Nilai']}"
    ws["C10"] = f"Pengantar Pencairan APBDES {info_df.loc[0,'Nilai']}"
    ws["C13"].value = (
    f"           Berdasarkan surat Kepala Desa {info_df.loc[0,'Nilai']} tanggal {info_df.loc[1,'Nilai']} "
    f"Nomor : {info_df.loc[2,'Nilai']} Perihal permohonan pengantar pencairan rekening kas desa, "
    f"maka bersama ini kami sampaikan kepada Cabang Pembantu Bank Jatim "
    f"Cabang Tuban di Rengel, untuk melakukan pencairan dana dari "
    f"Rekening Kas Desa, sesuai SPM sebagai berikut :"
)


    table_end = write_table(ws, table_df)
    after_total = write_total(ws, table_end)
    format_table(ws, table_end)
    write_signature(ws, after_total)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# STREAMLIT UI
# =========================
uploaded = st.file_uploader("Upload File Surat (Excel)", type=["xlsx"])

if uploaded:
    df = pd.read_excel(uploaded, header=None)
    nama, tgl, nomor = extract_info(df)

    table = extract_table(df, uploaded)

    info_df = pd.DataFrame({
        "Keterangan": ["Nama Desa","Tanggal","Nomor"],
        "Nilai": [nama, tgl, nomor]
    })

    st.table(info_df)
    st.dataframe(table)

    result = fill_to_template(info_df, table)

    st.download_button(
        "Download Surat Excel",
        data=result,
        file_name="Surat_Hasil_Ekstraksi.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
